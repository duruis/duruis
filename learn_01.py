# Immutable = you cannot change once you definite it
# str; int; float; bool; bytes; tuple

# Muttable = you can change them
# list []; set; dict

# If statements If statements If statements
"""
score = float(input("Score: "))
if score >= 90:
  print("Grade A")

elif score >= 80:
  print("Grade B")

elif score >= 70:
  print("Grade C")

elif score >= 60:
  print("Grade D")

else:
  print("You are fucked")
"""
"""
temperature = float(input("Temperature outside?: "))
if temperature > 35:
  print("Stay hydrated, it is hot out")
elif  temperature <= 25 and temperature >=18:
  print("It is nice out, get out of the house")
else:
  print("temperatures are dropping, bring a coat just in case")
"""
"""  
weight = float(input("Weight: "))
variable = input("(k)g or Pounds: ")
if variable.upper == "K":
  converted = weight / 0.45
  print("Weight in Pounds: " + str(converted))
else:
  converted = weight * 0.45
  print("Weight in Kilos: " + str(converted))
"""

# Lists Lists Lists
"""
list_1 = [1,2,3,4,5]
for i in list_1:
  print(i, end=" ")
   
list_1 = [1,2,3,4,5]
list_1.append(6)
list_1.remove (5)
list_1.insert(0,0)
list_1.pop(4)
list_1.reverse()
print(list_1)
List_1.clear
List_1.index
List_1.count
List_1.reverse
List_1.copy
list_2 = [20,30,40,50]
list_1.extend(list_2)
print(list_1)
"""
# List Comprehensions - concise way to create a list.
"""
square = []
for x in range (10):
  square.append(x**2)
print(square)
"""
"""
combs = []
for x in [1,2,3]:
 for y in [2,4,5,6]:
   if x != y:
     combs.append((x,y))
print(combs)
"""
"""
seasons = ['Spring', 'Summer', 'Autumn', 'Winter']
print(list(enumerate(seasons, start=1)))
"""
# Loops Loops Loops

"""
name = ["Edu", "Dudu", "Ed", "Eduardo"]
age = [23, 25, 29, 34]
for n, a in zip(name, age):
  print("{} is {} years old".format(n,a))
"""
"""
while True:
  print ("Who are you?")
  name = input()
  if name != "Edu":
    continue
  print("Hello Edu. What is your password? ")
  password = input ()
  if password == "Monkey":
    break
print("Access granted")
"""
"""
while True:
  try:
   x = float(input("What is x? "))
  except ValueError:
   print("x is not an integer")
  else:
   break
print(f"x is {x}")

"""
"""
while True:
  word = input("What is the magic word: ?")
  if word == "ed":
    print("You win") 
  else:
    print('Try again:')
"""
"""
i = 1
while i <= 5:
  print(i, end=" ")
  i = i + 1
"""
"""
count = 0
while count < 7:
  count += 1
  if count == 5:
    break
  print("you are an ass")
"""

# random random random
"""
import random
numbers = [1,2,3,4,5,6]
chance=random.choice(numbers)
print(chance)
"""
"""
import random
for i in range(1):
 print(random.randint(0,7))
"""

"""
import random
list = ('Ed', 'Edu', 'Ed', 'Edu')
pick=random.choice(list)
print(pick)
"""
"""
import random
for i in range(6):
 print(random.randint(0,61))
"""

"""
import random
def runGame():
  "the game code"
  while True:
   print("Game menu")
   print("1) Start a new game")
   print("2) Quit")
   
   choice = input("Enter your choice")
   if choice == "1":
     runGame()
   elif choice == "2":
     break 
"""
# Functions Functions Functions
"""
def great_name(First_name, Last_name):
  print(f"Hi {First_name} {Last_name}")
  print("Welcome to the show")
  
print("What's up everyone")
great_name("Ed", "Castro")
print("Hope you all have a great day")
"""
"""
def square(number):
  return number * number
print(square(4))
"""
"""
def main ():
 student = get_student()
 print(f"{student[0]} from {student[1]}")
 
def get_student():
  name = input("Name: ")
  house = input("House: ")
  return (name, house)

if __name__ == "__main__":
 main() 
"""

"""
def main ():
  x = float(input("What is x? "))
  print("x squared is", square (x))
  
def square(n):
  return pow (n, 3)

main()
"""

"""
def math (x,y=10):
  return x ** y
result = math(3)
print(result)
"""

"""
def function (*args, **kwargs ): # args and kwargs will run as many args u want and different types
 print(args, kwargs)
 pass
function (10, 34, 45, 90, c = True, d = 10)
"""
"""
def function (x, y, z = 10):
 print(x + y * z)
 pass
function (4, 6, )
"""
"""
def multiply(x, y):
  print(f"You called the multiply of x,y with the values of x = {x} and y= {y}")
  print(f"the multiply of x*y = {x*y}")
multiply(4,8)
"""
"""
def square (x):
  return x*x
print(square(9))
"""
"""
import random
magicnumber = random.randint(1,10)
while True:
 word = input("What is the magic number? ")
 if word == "quit":
   break

 try:
   playnumber = int(word)
 except ValueError:
   print("Try a different number without decimals")
   continue
    
 if playnumber == magicnumber:
  print('You got it right')
 elif playnumber < magicnumber:
  print("playnumber is lower")
 elif playnumber > magicnumber:
  print("Playnumber is higher")
  print("This is not the magic number, try again")
"""
"""
import random
numbers = random.choice (list(range(1, 61)))
print(numbers) 
"""

"""

import random
numbers = random.randint(0, 61)
print(numbers)
"""
"""
import random

cards = ["jack", "queen", "king"]
random.shuffle (cards)
for card in cards:
  print(card)
"""
"""
from random import shuffle
numbers = [1, 2, 3, 4, 5, 6, 7, 8, 9]
shuffle (numbers)
print (numbers)

from random import sample
names =['Ed','Edu', 'Dudu', 'Eduardo']
single_sample= sample(names,3) 
print(single_sample)
"""
# Class Class Class
"""
import random
class Dice:
 def roll(self):
   first = random.randint(1,6)
   second = random.randint(1,6)
   return first, second
   

Dice_01 = Dice()
print (Dice_01.roll())
"""

"""
class Point:
  def __init__(self, x,y ):
    self.x = x
    self.y = y
  
  def name():
    print("name")
  def draw():
    print("draw")
    
point_01=Point(10,20)
print(point_01.x)
print(point_01.y)
"""
"""
class Person:
   def __init__(self, name):
    self.name = name
  
   def talk(self):
    print(f"I am {self.name} how is your day going?")
    
Person_01 = Person ("Ed")
Person_01.talk()
Person_02 = Person ("Eduardo")
Person_02.talk()
"""
"""
class Person:
  def __init__(self,name):
    self.name = name
  
  def whoami(self):
    return "You are " + self.name

p1 = Person("Edu")
print(p1.whoami())
print(p1.name)
"""
"""
class human:
  def SayHello(self, name=None):
    if name != None:
      print("What's up", (name))
    else:
      print("Do I know you?")

object=human()
object.SayHello()
object.SayHello(" ")
"""
# Dictionary Dictionary Dictionary
"""
tel= {"ed": 3475, "edu": 3745, "dudu": 2049, "eduar": 2847}
tel["jo"] = 2103
del tel ["ed"]
print(tel)
"""
# looping
"""
knights = {"galand": "the pure", "robin": "the brave"}
for k, w, in knights.items():
  print(k,w)
"""
"""
questions = ["name", "quest", "favorite color"]
answer = ["lancelot", "the holy grail", "blue"]
for q,a in zip(questions, answer):
  print("What is your: {0} It is {1}".format(q,a))
"""
# using set to eliminate duplicates
"""
basket = ["apple", "orange", "apple", "papaia", "carrots"]
for f in sorted(set(basket)):
  print(f) 
"""

"""
b_year = int(input("What is your year of birth? "))
age = 2024 - b_year
print(age)
"""

"""
first = ("Edu")
last = ("monkey")
print(f"{first} is the coolest dude and his last name is {last}")
"""
"""
course = ("Python for beginners")
print(f"{course.replace("Python", "sex")}")
"""
"""
price_house = float(input("What is the price of the house? "))
good_bad_credit = input("Do you have good or bad credit? ").lower()
if good_bad_credit == "good":
  print(f"total_downpayment", {price_house * 0.1})
elif good_bad_credit == "bad":
 print("total_downpayment",   {price_house * 0.2})
print("Thanks for your business")
"""
"""
price = 100000
good_credit = False
if good_credit:
  down_payment = price * 0.1
else:
  down_payment = price * 0.2
  print(f"Down payment: " ${down_payment})
print("Thanks for your business")
"""
"""
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
  guess = int(input("Guess: "))
  guess_count += 1
  if guess == secret_number:
   print("You won")
   break
else:
  print("Sorry, you failed")
"""
"""
prices = [10,20,30,40]
for i in prices:
 print(f"You own me {sum(prices)}")
 break
"""
"""
list = [5,2,5,2,2]
for number in list:
  print("x" * number)
#or
list = [5,2,5,2,2]
for number in list:
  output= ""
  for first in range(number):
     output += "x"
  print(output)
"""
"""
matrix = [
[1, 2 , 3],
[4, 5, 6],
[7, 8, 9]
]
matrix [1] [0:3] = [10, 20, 30]
print(matrix) 
"""
"""
numbers = (1,2,4,3,5,6,7,1,2,3,7,8,9,5,3,2,1,4,5,6)
no_duplicates = list(set(numbers))
print(no_duplicates)
"""
"""
numbers = (0,1,3,4,5)
x,t,r,e,t = numbers
print((r*t)/e)
"""
"""
contact = {"name": "Edu",
  "age": "30",
  "phone": "10204",
  "city": "Nashive",
    }
print(contact ["age"])
"""
"""
phone = input("What is your fucking phone number bitch? ")
digital={"1": "one", "2": "two", "3": "three"}
for i in phone:
 print(digital.get(i), end=" ")
"""
# mkdir rmdir touch
"""
from pathlib import Path
path = Path()
for file in path.glob("*.py"):
 print(file)
"""
"""
from pathlib import Path
path = Path()
print(path.touch())
"""
"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference 

def process_workbook (filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    for row in range(2, sheet.max_row -1):
      cell=sheet.cell(row,3)
      corrected_price = cell.value * 0.9
      corrected_price_cell = sheet.cell(row, 4)
      corrected_price_cell.value=corrected_price

    values = Reference (sheet,
    min_value=2,
    max_value=sheet.max_row -1,
    min_col=4,
    max_col=4)
            
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")

      
    wb.save(filename)
  """

#################################################
'''x = [i for i in range (10)]
y = [i for i in range (10) if i % 2 == 0]
print(x + y)

def function (x, y, z = 10):
 print(x + y * z)
 pass
function (4, 6, )

# args and kwargs will run as many args u want and different types
def function (*args, **kwargs ): 
 print(args, kwargs)
 pass
function (10, 34, 45, 90, c = true, d = 10)

# using * and ** 
def function ( a, b, c= True, d= False):
  print(a, b, c, d)
  pass
function (*[1, 2], **{"c":  "cool", "d": "nice"})
'''
'''
name = input ("What's your name? ")
print("Hello " + name)'''
"""
bday = input("When is your birthday year? ") 
age = 2024 - int(bday)
print(age)
"""
"""
First = float(input("First: "))
Second = float(input("Second: "))
Sum = First + Second
print("Sum:", + Sum)
"""
"""
animal = (1223, 1223, 1234, 2323, 3445, 4323, 23, 23, )
print(animal.count(23))
"""
"""
animal = 230
print(animal <= 240 and animal > 210)
"""
"""
temperature = float(input("Temperature: " ))
if temperature > 20 and temperature <= 35:
 print("Lets stay home, and have sex")
 print("Enjoy the day, smoke weed, have sex again lol")

print("Done")
"""
"""
temperature = float(input("Temperature outside?: "))
if temperature > 35:
  print("Stay hydrated, it is hot out")
elif  temperature <= 25 and temperature >=18:
  print("It is nice out, get out of the house")

else:
  print("temperatures are dropping, bring a coat just in case")
  
"""
"""
weight = float(input("Weight: "))
variable = input("(k)g or Pounds: ")
if variable.upper == "K":
  converted = weight / 0.45
  print("Weight in Pounds: " + str(converted))
else:
  converted = weight * 0.45
  print("Weight in Kilos: " + str(converted))

"""
"""
i = 1
while i <= 5:
  print(i)
  i = i + 1
"""
"""
#list of names
names_list = ["Edu", "Ed", "ed", "Du", "Dudu", "Eduardo"]
#Function
def find_name(name, names_list):
  if name in names_list: 
    return f"{name} is in the list."
  else:
    return f"{name} is not listed."
 
#Input= name to find 
name_to_find = input ("Enter the name you want to find: ")

#Result
result = find_name(name_to_find, names_list)
print(result)
"""
"""
numbers = [2, 4, 6, 8, 10]
for item in numbers:
  print(item)
"""
"""
numbers = range (0, 60, 2)
for i in numbers:
  print(i)
"""
"""  
color = input("What is your favorite color: ?")
print(f"I also love {color}")

"""
"""
print("What is your name? ")
name = input()
print ("It is nice to meet you, {}" .format (name)) 

"""
""" 
while True:
  print ("Who are you?")
  name = input()
  if name != "Edu":
    continue
  print("Hello Edu. What is your password? ")
  password = input ()
  if password == "Monkey":
    break
print("Access granted")
"""
"""
for i in [ 1, 2, 3, 4, 5, 6, 7, 8, 9, ]:
  if i == 11:
    break
else:
  print("only executed no item in the list in equal to 3")
"""
"""
name = ["Edu", "Dudu", "Ed", "Eduardo"]
age = [23, 25, 29, 34]
for n, a in zip(name, age):
  print("{} is {} years old".format(n,a))
"""
"""
print ("Hello world\nIt is been a freaking beautiful journey\nThanks God for another beautiful day")
"""
"""
def add(x, y):
 return x + y
result = add(23, 77)
print(result)

"""
"""
x = float(input("What is x? "))
y = float(input("What is y? "))
print(f"{x + y: ,}")
"""
"""
x = float(input("What is x? "))
y = float(input("What is y? "))
print(round(x / y, 2))
"""
"""
def hello(to="World"):
   print("hello,", to)
   
hello()
name = input("What is your name? ")
hello (name)
"""
"""
def main ():
  x = float(input("What is x? "))
  print("x squared is", square (x))
  
def square(n):
  return pow (n, 2)

main()
"""
""" 
score = float(input("Score: "))
if score >= 90:
  print("Grade A")

elif score >= 80:
  print("Grade B")

elif score >= 70:
  print("Grade C")

elif score >= 60:
  print("Grade D")

else:
  print("You are fucked")
"""
"""
def main ():
  x = float(input("What is x? "))
  if is_even(x):
    print("Even")
  else:
    print("Odd")
def is_even (n):
    return n % 2 == 0
     
main ()
"""
"""
name = input("What is your name? ")
match name:
  case "Edu" | "Ed" | "Dudu":
    print ("Pira")
  case "Mel" | "Me" | "Meme":
    print("Recife")
  """
"""  
i = 5
while i != 0:
  print("meow")
  i = i - 1
"""
"""
for i in range (20):
  print ("meow")
  """
"""
print ("meow\n" * 3, end="")  
  """
""" 
while True:
  n = int(input("What is n? "))
  if n > 0:
     break
for _ in range(n):
  print("meow")
"""
"""
students = ["Edu", "Dudu", "Ed", ]
for i in range(len(students)):
  print(i + 1, students[i])
  
"""
"""
animals = [
  {"ID": "PUSP", "YEAR": "TF", "SEX": "M"},
  {"ID": "PUSP", "YEAR": "TT", "SEX": "F"},
  {"ID": "PUSP", "YEAR": "TTT", "SEX": "M"},
]  
for individual in animals:
 print(individual ["ID"], individual ["YEAR"], individual ["SEX"], sep=(", "))
"""
"""
def main ():
 print_row(4)
  
def print_row (width):
  print("?" * width)
  
main()
"""
"""
def main():
  print_square(3)
  
def print_square(n):
  for i in range(n):  
   print("#" * n)

main ()
"""
"""
while True:
  try:
   x = float(input("What is x? "))
  except ValueError:
   print("x is not an integer")
  else:
   break
print(f"x is {x}")

"""
"""
import random
numbers = random.choice (list(range(61)))
print(numbers)  

"""
"""
import random
numbers = random.randint(0, 61)
print(numbers)

"""
"""
import random

cards = ["jack", "queen", "king"]
random.shuffle (cards)
for card in cards:
  print(card)
"""
"""
import statistics
print(statistics.mean ([10, 20]))
"""
"""
import sys
if len(sys.argv) < 2:
   print ("Too few arguments")
elif len(sys.argv) > 2:
  print ("Too many arguments")
else:
 print("What's up? My name is", sys.argv[1])
"""
"""
def main():
  hello("world")
  goodbye("world")
  
def hello(name):
  print(f"hello, {name}")
  
def goodbye(name):
  print(f"goodbye, {name}")

if __name__ == "__main__":
 main ()

# continue ...create a new file

import sys
from learn_01 import hello
if len(sys.argv) == 2:
   hello (sys.argv [1])

"""
"""
def main ():
  x = float(input("What is x? "))
  print("x square is ", square(x))
  
def square(n):
  return n * n

if __name__ == "__main__":
 main ()
 
#in a new file

from learn_01 import square

def test_positive():
  assert square(2) == 4
  assert square(3) == 9

def test_negative():
  assert square(-2) == 4
  assert square(-3) == 9

def test_zero():
  assert square(0) == 0
"""
"""
name = input("What is your name? ")
with open("names.txt", "a") as file:
 file.write(f"{name}\n")

with open("names.txt", "r") as file:
  for line in file:
    print("hello,", line.rstrip())

"""
"""
names = []
with open("names.txt") as file:
  for line in file:
    names.append(line.rstrip())
for name in sorted(names):
 print(f"hello, {name}") 

"""
"""
with open ("names.txt") as file:
   for line in file:
     name, location = line.rstrip().split(",")
     print(f"{name} is in {location}")
"""
"""
#to read and sort through the csv file
 import csv
 names = []
 with open ("names.txt") as file:
   reader = csv.DictReader(file)
   for name , location in reader:
     names.append({"name":name, "location":location})

 print(f"{name} is in {location}")
"""
"""
import csv
name = input("What is your name? ")
home = input("Where is your home? ")

with open("students.csv", "a") as file:
  writer = csv.DictWriter(file, fieldnames=["name", "home"])
  writer.writerow({"name": name, "home": home})
  
"""
"""

import sys
from PIL import Image
novo = []

for arg in sys.argv[1:]:
  image = Image.open(arg)
  novo.append(image)
  
novo[0].save(
  "new_file.gif", save_all=True, append_novo=[novo[1]], duration=100, loop=0) 

"""
"""
email = input("What is your email?").strip()

if "a" in email and  "." in email: 
  print("Valid")
else:
  print("Invalid")   
  
"""
"""
import re
name = input("What is your name? ").strip()
if matches := re.search(r"^(.+), *(.+)$", name):
  name = matches.group(1) + " " + matches.group(2)
  print(f"hello, {name}")

"""


"""
name = input("Name: ?")
house = input("House: ?")
print(f"{name} from {house}")
"""
"""
def main ():
 student = get_student()
 print(f"{student[0]} from {student[1]}")
 
def get_student():
  name = input("Name: ")
  house = input("House: ")
  return (name, house)

if __name__ == "__main__":
 main() 
 
"""
'''
class Student:
    def __init__(self, name, house):
        if not name:
            raise ValueError("Missing name")
        if house not in ["Pira", "Recife", "Olimpia"]:
            raise ValueError("Invalid house")
        self.name = name
        self.house = house

    def __str__(self):
        return f"{self.name} from {self.house}"


def get_student():
    name = input("Name: ")
    house = input("House: ")
    return Student(name, house)


def main():
    student = get_student()
    print(student)


if __name__ == "__main__":
    main()
'''
